# -*- coding: utf-8 -*-

import openpyxl as opxl
import numpy as np


def read_excel(file_path):
    #
    wb = opxl.load_workbook(file_path)
    ws = wb.active
    #
    data_range = "A1:N" + str(ws.max_row)
    #
    data = []
    for line_data in ws[data_range]:
        tmp_data = [cell.value for cell in line_data]
        data.append(tmp_data)
    #
    return np.array(data)
