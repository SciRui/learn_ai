import os
import openpyxl as opxl
import numpy as np

def read_data(in_file_path):
    #
    data = []
    wb = opxl.load_workbook(in_file_path)
    ws = wb["data"]
    for line_data in ws["B2:ABY58"]:
        tmp_data = [cell.value for cell in line_data]
        data.append(tmp_data)
    return np.array(data, dtype=np.float32)

def one_hot_labels(labels):
    train_labels = []
    for label in labels:
        one_hot_labels = [0,0]
        if label == 0:
            one_hot_labels[0] = 1
        else:
            one_hot_labels[1] = 1
        train_labels.append(one_hot_labels)
    return np.array(train_labels)

def train_data_iterator(samples, labels, iteration_steps, chunkSize):
    """
    Iterator/Generator: get a batch of data
    这个函数是一个迭代器/生成器，用于每一次只得到 chunkSize 这么多的数据
    用于 for loop， just like range() function
    """
    if len(samples) != len(labels):
        raise Exception('Length of samples and labels must equal')
    stepStart = 0  # initial step
    i = 0
    while i < iteration_steps:
        stepStart = (i * chunkSize) % (labels.shape[0] - chunkSize)
        yield i, samples[stepStart:stepStart + chunkSize], labels[stepStart:stepStart + chunkSize]
        i += 1


def test_data_iterator(samples, labels, chunkSize):
    """
    Iterator/Generator: get a batch of data
    这个函数是一个迭代器/生成器，用于每一次只得到 chunkSize 这么多的数据
    用于 for loop， just like range() function
    """
    if len(samples) != len(labels):
        raise Exception('Length of samples and labels must equal')
    stepStart = 0  # initial step
    i = 0
    while stepStart < len(samples):
        stepEnd = stepStart + chunkSize
        if stepEnd < len(samples):
            yield i, samples[stepStart:stepEnd], labels[stepStart:stepEnd]
            i += 1
        stepStart = stepEnd

def create_data(data, save_data = True, save_path = "I:/learn_tf/Data/water_quality/new_water_data.xlsx"):
    #
    row,col = data.shape
    new_data = np.ndarray([row*20,col])
    k = 0
    for i in range(row):
        new_data[k,:]=data[i,:]
        tmp_data = np.random.random([19,col])*0.001 + data[i,:]
        tmp_data[:,-1]=data[i,-1]
        new_data[k+1:k+20,:] = tmp_data
        k += 20
    if save_data:
        wb = opxl.Workbook()
        ws = wb.create_sheet("data",index=0)
        for i in range(new_data.shape[0]):
            ws.append(list(new_data[i,:]))
        wb.save(save_path)
    return new_data

def get_cnn_data(data):
    #
    return data.reshape([data.shape[0],1,1,data.shape[1]])

current_dir = os.getcwd()
data = read_data(current_dir + r"\Data\water_quality\water_data.xlsx")
bad_data_col = np.unique(np.argwhere(data<0)[:,1])
data = np.delete(data, bad_data_col, axis=1)
data = create_data(data,save_data = False)
np.random.shuffle(data)  

split_index = int(np.around(0.7*data.shape[0]))
#
ml_samples = data[:,:-1]
ml_labels = data[:,-1]
#
train_samples = data[:split_index, :-1]
test_samples = data[split_index:, :-1]
train_images = get_cnn_data(train_samples)
test_images = get_cnn_data(test_samples)
ml_train_labels = np.array(data[:split_index, -1])
dl_train_labels = one_hot_labels(data[:split_index, -1])
ml_test_labels = np.array(data[split_index:, -1])
dl_test_labels = one_hot_labels(data[split_index:, -1])

# print(test_samples)
# print(test_labels)
