# -*- coding: utf-8 -*-

import sys

sys.path.append("..\\")

import openpyxl as opxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import randint as scipy_randint
from scipy.stats import expon as scipy_expon
from sklearn import svm
from sklearn import model_selection
from sklearn import metrics
from sklearn import utils
from  metrics_lib import metrics_regression
from  metrics_lib import data_similarity
from plot_lib import plot_metrics
from plot_lib import plot_data
from preprocess_lib import data_split


def read_excel(file_path):
    #
    wb = opxl.load_workbook(file_path)
    ws = wb.active
    #
    data_range = "A1:N" + str(ws.max_row)
    #
    data = []
    for line_data in ws[data_range]:
        tmp_data = [cell.value for cell in line_data]
        data.append(tmp_data)
    #
    return np.array(data)


def init_model():
    #
    init_model = svm.LinearSVR()
    #
    return init_model

def train_cv_model(init_model, X, y, n_splits = 10, training_size = 0.7, test_size = 0.3, search = "Grid"):
    #
    cv, indices, cos_theta = data_split.imitate_split(y, n_splits, training_size, test_size, cos_theta_lim = 0.8)
    print("cos_theta_list:",cos_theta)
    #
##    cv = model_selection.ShuffleSplit(n_splits, training_size, test_size,  random_state = 0)
    #
    train_val_data = []
    train_val_scores = []
    train_val_params = []
    model = None
    #
    if search == "Grid":
        optimize_parameters = {"C":[1.,5.,10.,20.,25,35.,40.,50.],
                               "epsilon":[0.001,0.005,0.1,0.3,0.5,0.7,0.9],
                               "loss":["epsilon_insensitive","squared_epsilon_insensitive"],
                               "max_iter":[200,500,1000,1400,1700,2000,2500,3000],
                               "random_state":[0,1,5,10,20,30,50,60,80,100]}
        #
        model = model_selection.GridSearchCV(init_model,
                                             optimize_parameters,
                                             cv = cv,
                                             refit = "r2",
                                             scoring=("r2","neg_mean_squared_error"),
                                             return_train_score = True,
                                             n_jobs = 4)
    elif search == "Random":
        #
        optimize_parameters = {"C":scipy_expon(scale = 20.0),
                               "epsilon":scipy_expon(scale = 0.005),
                               "loss":["epsilon_insensitive","squared_epsilon_insensitive"],
                               "max_iter":scipy_randint(200,3000),
                               "random_state":scipy_randint(0,100)}
        #
        model = model_selection.RandomizedSearchCV(init_model,
                                                   optimize_parameters,
                                                   refit = "r2",
                                                   scoring=("r2","neg_mean_squared_error"),
                                                   cv = cv,
                                                   n_iter = 100,
                                                   return_train_score = True,
                                                   n_jobs = 4)
    else:
        raise("""Error Parameter input "search"!""")
        return model, train_val_scores, train_val_data, train_val_params
    #
    model.fit(X, y)
    ###
    #
    training_r2 = model.cv_results_.get("mean_train_r2")
    val_r2 = model.cv_results_.get("mean_test_r2")
    training_neg_mse = model.cv_results_.get("mean_train_neg_mean_squared_error")
    val_neg_mse = model.cv_results_.get("mean_test_neg_mean_squared_error")
    train_val_scores.extend([training_r2, val_r2, training_neg_mse, val_neg_mse])
    ###
    #
    X_train, X_val = X[indices[0]], X[indices[1]]
    y_train, y_val = y[indices[0]], y[indices[1]]
    prediction_train = model.predict(X_train)
    prediction_val = model.predict(X_val)
    train_val_data.extend([X_train, X_val, y_train, y_val, prediction_train, prediction_val])
    ###
    if search == "Grid":
        train_val_params.extend([optimize_parameters, model.best_params_])
    elif search == "Random":
        tmp_params = {}
        params = model.cv_results_.get("params")
        keys = list(params[0].keys())
        for key in keys:
            tmp_params.update({key:[]})

        for params_dict in params:
            for key in keys:
                tmp_params.get(key).append(params_dict.get(key))

        for key in keys:
            tmp_params.get(key).sort()
            
        train_val_params.extend([tmp_params, model.best_params_])
    #
    return model, train_val_scores, train_val_data, train_val_params

def test_model(X, y):
    #
    model_info = []
    #
    predictions = model.predict(X)
    r2 = metrics.r2_score(y, predictions)
    mse = metrics.mean_squared_error(y, predictions)
    slope, intercept = metrics_regression.fitting_line_params(y, predictions)
    model_info.extend([r2, mse, slope, intercept])
    #
    return predictions, model_info
    

if __name__ == "__main__":
    #
    training_set = read_excel(r"E:\About_Program\Python\Projects\learn_AI\Data\sklearn_data\training_sklearn_boston.xlsx")
    validation_set = read_excel(r"E:\About_Program\Python\Projects\learn_AI\Data\sklearn_data\validation_sklearn_boston.xlsx")
####    test_set = read_excel(r"E:\About_Program\Python\Projects\learn_AI\Data\sklearn_data\test_sklearn_boston.xlsx")
    ##
    #
    training_cv_set = np.vstack((training_set, validation_set))
    X_training_cv, y_training_cv = training_cv_set[:,:-1], training_cv_set[:,-1]
    #
    init_model = init_model()
    model, train_val_data, train_val_scores,train_val_params = train_cv_model(init_model, X_training_cv, y_training_cv,search = "Random",
                                                     n_splits = 1, training_size = 0.8, test_size = 0.2)
    #
    plot_metrics.plot_learning_curve1(model,X_training_cv, y_training_cv, n_splits = 5, scoring = "r2")
##    plot_metrics.plot_learning_curve(model, X_training_cv, y_training_cv, 5, train_sizes=(0.2,0.35,0.5,0.65,0.8,0.95,1.0), scoring="r2")
##    plot_metrics.plot_valdation_curve(init_model, X_training_cv, y_training_cv, 5, train_val_params, scoring="r2")
##    plot_metrics.plot_model_score(train_val_data, train_val_scores)



